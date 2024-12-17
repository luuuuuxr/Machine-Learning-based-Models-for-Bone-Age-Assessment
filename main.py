import time
import matplotlib.pyplot as plt
import seaborn as sns
import numpy as np
import pandas as pd
import shap
from openpyxl.reader.excel import load_workbook
from sklearn.ensemble import RandomForestRegressor, AdaBoostRegressor
from sklearn.linear_model import Lasso
from sklearn.metrics import mean_absolute_error, mean_squared_error, r2_score, explained_variance_score
from sklearn.model_selection import KFold, cross_val_predict
from sklearn.neighbors import KNeighborsRegressor
from sklearn.neural_network import MLPRegressor
from sklearn.svm import SVR
from sklearn.tree import DecisionTreeRegressor
from xgboost import XGBRegressor
from selenium import webdriver
from matplotlib import image as mpimg
from math import sqrt

# SHAP初始化
shap.initjs()

# 统一的图形参数
plt.rcParams['font.sans-serif'] = ['Microsoft YaHei']  # 显示中文
plt.rcParams['axes.unicode_minus'] = False  # 显示负号


# ==================== 初始化模型 ====================
def init_models():
    return {
        'DecisionTree': DecisionTreeRegressor(random_state=31, max_depth=4, min_samples_split=19),
        'RandomForest': RandomForestRegressor(random_state=4832, n_estimators=99, max_depth=11),
        'KNN': KNeighborsRegressor(n_neighbors=11),
        'Lasso': Lasso(alpha=0.000270405, max_iter=1000000),
        'SVR': SVR(kernel='linear', C=1.0, epsilon=0.1),
        'ANN': MLPRegressor(hidden_layer_sizes=(100,), activation='tanh', solver='adam',
                            learning_rate='adaptive', max_iter=1000)
    }


# ==================== 模型交叉验证及可视化 ====================
def cross_validate_and_plot(models, X, Y, n_splits=5):
    kf = KFold(n_splits=n_splits, shuffle=True, random_state=42)
    fig, axes = plt.subplots(4, 2, figsize=(15, 20))
    axes = axes.flatten()

    for i, (model_name, model) in enumerate(models.items()):
        y_pred = cross_val_predict(model, X, Y.ravel(), cv=kf)
        y_pred = np.round(y_pred, decimals=2)
        mse = mean_squared_error(Y, y_pred)
        r2 = r2_score(Y, y_pred)

        # 绘制拟合曲线和置信区间
        sns.regplot(x=Y, y=y_pred, ax=axes[i], scatter_kws={"s": 10}, line_kws={"color": "red"})
        axes[i].fill_between(Y.ravel(), y_pred - 1.96 * np.sqrt(mse), y_pred + 1.96 * np.sqrt(mse), color='red',
                             alpha=0.2)
        axes[i].set_title(f'{model_name}\nMSE: {mse:.4f}, R2: {r2:.4f}')
        axes[i].set_xlabel('True Values')
        axes[i].set_ylabel('Predictions')
        axes[i].grid(True)

    plt.tight_layout()
    plt.savefig('cross_validation_plot.png')
    plt.show()


def plot_jointgrid_multiple(y_train, y_train_preds, y_test, y_test_preds, model_names):
    # 创建一个2x3的网格来展示六个模型的图
    fig, axes = plt.subplots(2, 3, figsize=(15, 8), dpi=150)
    axes = axes.flatten()  # 将轴阵列展平
    platte = ['#F6D8C2', '#C5D3E8']
    platte2 = ['#D22225', '#25307A']

    for i, model_name in enumerate(model_names):
        ax = axes[i]
        # 创建包含训练集和测试集真实值与预测值的数据框
        data_train = pd.DataFrame({
            'True': y_train.ravel(),
            'Predicted': y_train_preds[model_name].ravel(),
            'Data Set': 'Train'
        })

        data_test = pd.DataFrame({
            'True': y_test.ravel(),
            'Predicted': y_test_preds[model_name].ravel(),
            'Data Set': 'Test'
        })

        data = pd.concat([data_train, data_test])

        # 创建散点图
        sns.scatterplot(data=data, x="True", y="Predicted", hue="Data Set", palette=platte, alpha=0.8,ax=ax)

        # 添加训练集和测试集的回归线
        sns.regplot(data=data_train, x="True", y="Predicted", scatter=False, color=platte2[0], ax=ax,
                    label='Train Regression')
        sns.regplot(data=data_test, x="True", y="Predicted", scatter=False, color=platte2[1], ax=ax,
                    label='Test Regression')

        # 添加拟合优度
        r2_train = r2_score(y_train, y_train_preds[model_name])
        r2_test = r2_score(y_test, y_test_preds[model_name])

        # 显示 R2 值
        ax.text(0.95, 0.1, f'Train $R^2$ = {r2_train:.3f}', transform=ax.transAxes, fontsize=12,
                verticalalignment='bottom', horizontalalignment='right')
        ax.text(0.95, 0.05, f'Test $R^2$ = {r2_test:.3f}', transform=ax.transAxes, fontsize=12,
                verticalalignment='bottom', horizontalalignment='right')

        # 添加模型名称
        ax.text(0.85, 0.25, f'{model_name}', transform=ax.transAxes, fontsize=14,
                verticalalignment='bottom', horizontalalignment='right')

        # 中心线
        ax.plot([data['True'].min(), data['True'].max()], [data['True'].min(), data['True'].max()], c="black",
                alpha=0.5, linestyle='--')

        # 添加图例
        ax.legend()

    # 调整布局以确保所有图不重叠
    plt.tight_layout()
    plt.savefig("combined_six_models_plot.pdf", format='pdf', bbox_inches='tight')
    plt.savefig("combined_six_models_plot.png", bbox_inches='tight')
    plt.show()


# ==================== 模型训练和评估 ====================
def train_and_evaluate_all(models, X_train, Y_train, X_test):
    Y_train_preds = {}
    Y_test_preds = {}
    for model_name, model in models.items():
        model.fit(X_train, Y_train.ravel())
        y_train_pred = model.predict(X_train)
        y_test_pred = model.predict(X_test)
        y_train_pred = np.round(y_train_pred, decimals=2)
        y_test_pred = np.round(y_test_pred, decimals=2)
        Y_train_preds[model_name] = y_train_pred
        Y_test_preds[model_name] = y_test_pred

    return Y_train_preds, Y_test_preds


def evaluation(y_test, predictions):
    for i, (model_name, y_pred) in enumerate(predictions.items()):
        difference = y_test.ravel() - y_pred
        where = np.where(abs(difference) <= 1.0, 1, 0)
        cor = np.sum(where == 1)

        print('-------------------------------------------')
        print(model_name)
        # print("测试集数据:{}".format(y_test))
        # print("预测的结果:{}".format(y_pred))
        # print("绝对差值 :{}".format(difference))
        # print("筛选后的数据:{}".format(where))
        print("正确的个数:{}".format(cor))
        print("总个数:{}".format(len(y_test)))
        print("手动准确度:{}%".format(round(cor / len(y_test) * 100, 2)))

        # 越小越好，趋于0完美
        print("平均绝对误差MAE:{}".format(round(mean_absolute_error(y_test, y_pred), 4)))
        print("平均相对误差MRE:{}".format(round(np.mean(np.abs(y_test - y_pred) / y_test), 4)))
        print("均方误差MSE:{}".format(round(mean_squared_error(y_test, y_pred), 4)))
        print("均方根误差RMSE:{}".format(round(sqrt(mean_squared_error(y_test, y_pred)), 4)))

        # 越大越好，趋于1完美
        print("决定系数R2:{}".format(round(r2_score(y_test, y_pred), 4)))
        print("可解释变异EV:{}".format(round(explained_variance_score(y_test, y_pred), 4)))


# ==================== 可视化预测结果 ====================
def plot_all_predictions(Y_test, predictions):
    fig, axes = plt.subplots(4, 2, figsize=(15, 15))
    axes = axes.flatten()
    for i, (model_name, y_pred) in enumerate(predictions.items()):
        ax = axes[i]
        ax.scatter(Y_test, y_pred)
        ax.plot([Y_test.min(), Y_test.max()], [Y_test.min(), Y_test.max()], 'k--', lw=4)
        ax.set_title(f"{model_name}")
        ax.set_xlabel('True Values')
        ax.set_ylabel('Predictions')
        ax.grid(True)

    plt.tight_layout()
    plt.savefig('predict.png')
    plt.show()


def plot_bland_altman(Y_test, Y_test_preds):
    fig, axes = plt.subplots(4, 2, figsize=(15, 15))
    axes = axes.flatten()
    for i, (model_name, y_pred) in enumerate(Y_test_preds.items()):
        ax = axes[i]
        mean = np.mean([Y_test.ravel(), y_pred.ravel()], axis=0)
        diff = Y_test.ravel() - y_pred.ravel()
        md = np.mean(diff)
        sd = np.std(diff, axis=0)
        ax.scatter(np.hstack(mean), np.hstack(diff))
        ax.axhline(md, color='blue', linestyle='--')
        ax.axhline(md + 1.96 * sd, color='red', linestyle='--')
        ax.axhline(md - 1.96 * sd, color='red', linestyle='--')
        ax.set_title(f"{model_name}")
        ax.set_xlabel('Mean')
        ax.set_ylabel('Difference')
        ax.set_ylim(-2, 2)
        ax.grid(True)
    plt.tight_layout()
    plt.savefig('Bland_altman.png')
    plt.show()


def plot_boxplot(Y_test, predictions):
    data = {'Model': [], 'Residuals': []}

    for model_name, y_pred in predictions.items():
        residuals = Y_test.ravel() - y_pred.ravel()
        data['Model'].extend([model_name] * len(residuals))
        data['Residuals'].extend(residuals.tolist())

    df = pd.DataFrame(data)

    # Ensure 'Model' column is of type str
    df['Model'] = df['Model'].astype(str)

    # Ensure 'Residuals' column is numeric
    df['Residuals'] = pd.to_numeric(df['Residuals'], errors='coerce')

    # Print a sample of the dataframe to check data types and values
    print(df.head())
    print(df[df['Model'] == 'DecisionTree'].head())  # Check residuals data for a specific model

    plt.figure(figsize=(14, 7))

    # Define a palette with different colors for each model
    palette = sns.color_palette("dark", len(df['Model'].unique()))

    # Boxplot and Stripplot for residuals
    sns.boxplot(x='Model', y='Residuals', data=df, color='lightgray', showfliers=False)
    sns.stripplot(x='Model', y='Residuals', data=df, jitter=True, palette=palette, edgecolor='black', size=8,
                  linewidth=1, alpha=0.8)

    plt.title('Accuracy of Predictions')
    plt.ylabel('Accuracy')
    plt.legend(title='Model', bbox_to_anchor=(1.05, 1), loc='upper left')
    plt.grid(True)

    plt.tight_layout()
    plt.savefig('boxplot.png')
    plt.show()


# ==================== 预测结果保存 ====================
def save_all_predictions_to_excel(original_file, predictions_file, sheet_name, Y_train, Y_train_preds, Y_test,
                                  Y_test_preds):
    # 读取原始Excel文件
    book = load_workbook(original_file)
    writer = pd.ExcelWriter(original_file, engine='openpyxl')
    writer.book = book
    writer.sheets = {ws.title: ws for ws in book.worksheets}

    # 读取原始数据
    original_df = pd.read_excel(predictions_file, sheet_name=sheet_name, engine='openpyxl')

    # 将训练预测结果与原始数据合并
    for model_name, y_pred in Y_train_preds.items():
        # 过滤出有效的索引，即非NaN的索引
        valid_index = Y_train['Original_Index'].dropna()
        valid_y_pred = y_pred[valid_index.index]  # 确保预测值与有效索引对齐
        original_df.loc[valid_index, model_name] = valid_y_pred

    # 将测试预测结果与原始数据合并
    for model_name, y_pred in Y_test_preds.items():
        valid_index = Y_test['Original_Index'].dropna()
        valid_y_pred = y_pred[valid_index.index]  # 确保预测值与有效索引对齐
        original_df.loc[valid_index, model_name] = valid_y_pred

    # 将合并后的数据写回原始工作表
    # 将合并后的数据写回原始工作表
    original_df.to_excel(writer, sheet_name=sheet_name, index=False)

    # 保存工作簿
    writer.save()
    writer.close()


def save_test_predictions_to_excel(original_file, predictions_file, sheet_name, Y_test, Y_test_preds):
    # 读取原始Excel文件
    book = load_workbook(original_file)
    writer = pd.ExcelWriter(original_file, engine='openpyxl')
    writer.book = book
    writer.sheets = {ws.title: ws for ws in book.worksheets}

    # 读取原始数据
    original_df = pd.read_excel(predictions_file, sheet_name=sheet_name, engine='openpyxl')

    # 将测试预测结果与原始数据合并
    for model_name, y_pred in Y_test_preds.items():
        valid_index = Y_test['Original_Index'].dropna()
        valid_y_pred = y_pred[valid_index.index]  # 确保预测值与有效索引对齐
        original_df.loc[valid_index, model_name] = valid_y_pred

    # 将合并后的数据写回原始工作表
    # 将合并后的数据写回原始工作表
    original_df.to_excel(writer, sheet_name=sheet_name, index=False)

    # 保存工作簿
    writer.save()
    writer.close()



# ==================== 机器学习可解释性 ====================
# 获取力图
def capture_shap_force_plot(html_file_path, output_file_path):
    # 使用 webdriver-manager 自动处理 ChromeDriver
    driver = webdriver.Chrome()

    # 打开HTML文件
    url = "file:///" + 'E:/02  硕士/04  实验室工作/01  X射线估计骨龄/04  寒假任务/0617  任务/' + html_file_path
    driver.get(url)
    driver.set_window_size(1800, 300)  # 设置浏览器窗口的大小为1920x1080
    # 给予JavaScript足够的时间执行
    time.sleep(20)  # 可能需要调整等待时间
    # 获取截图
    driver.save_screenshot(output_file_path)

    if html_file_path == 'force_plot_7.html':
        # 关闭浏览器
        driver.quit()


def generate_force_plot(i, shap_values, axs_force, model_name):
    # 生成html
    shap_html = shap.force_plot(shap_values[0], show=False, matplotlib=False)
    html_path = f'force_plot_{i}.html'
    img_path = f'force_plot_{i}.png'
    shap.save_html(html_path, shap_html)

    capture_shap_force_plot(html_path, img_path)

    img = mpimg.imread(img_path)
    axs_force[i].imshow(img)
    axs_force[i].axis('off')  # 不显示坐标轴
    axs_force[i].set_title(f'{model_name}')


def plot_shap_summary(shap_values_dict, features):
    # 将数据转换为DataFrame
    df = pd.DataFrame(shap_values_dict, index=features).abs()

    colors = ['#a51890', '#D22225', '#f85a40', '#10763F', '#ffc845', '#037ef3']

    # 计算每个特征的总重要性，用于排序
    df['Total'] = df.sum(axis=1)
    df = df.sort_values(by='Total', ascending=True).drop(columns='Total')

    # 绘制堆叠条形图
    ax = df.plot(kind='barh', stacked=True, figsize=(10, 6), color=colors)
    ax.set_xlabel('Rank Score of Importance')
    ax.set_ylabel('Features')
    ax.set_title('Feature Importance Across Models')
    plt.savefig('Feature Importance')
    plt.show()


def generate_shap_plots(models, X_train, Y_train, X_test, Y_test):
    shap_values_dict = {}
    fig_force, axs_force = plt.subplots(6, 1, figsize=(20, 20))
    fig_waterfall, axs_waterfall = plt.subplots(2, 3, figsize=(12, 20))
    fig_summary, axs_summary = plt.subplots(2, 3)
    fig_summary_bar, axs_summary_bar = plt.subplots(2, 3)

    axes_force = axs_force.flatten()
    axes_waterfall = axs_waterfall.flatten()
    axes_summary = axs_summary.flatten()
    axes_summary_bar = axs_summary_bar.flatten()

    for i, (model_name, model) in enumerate(models.items()):
        model.fit(X_train, Y_train.ravel())

        if model_name in ['DecisionTree', 'RandomForest', 'XGB']:
            explainer = shap.TreeExplainer(model)
        else:
            explainer = shap.KernelExplainer(model.predict, shap.kmeans(X_test, k=5))
        shap_values = explainer(X_test)
        shap_values_dict[model_name] = shap_values.values.mean(axis=0)  # 存储每个特征的平均绝对SHAP值

        # 力图
        # generate_force_plot(i, shap_values, axes_force, model_name)

        # 假定 shap_values 是一个包含 shap.Explanation 对象的列表
        first_shap_value = shap_values[0]
        # 检查 base_values 并转换为标量
        if isinstance(first_shap_value.base_values, np.ndarray) and first_shap_value.base_values.size == 1:
            first_shap_value.base_values = first_shap_value.base_values.item()
        # 水坽图
        plt.sca(axes_waterfall[i])
        shap.plots.waterfall(first_shap_value, show=False, max_display=8)  # 使用修改后的 Explanation 对象
        axes_waterfall[i].set_title(f'{model_name}')
        plt.xticks(fontsize=8)
        plt.yticks(fontsize=8)

        # 特征总结图
        plt.sca(axes_summary[i])
        shap.summary_plot(shap_values, X_test, show=False)
        axes_summary[i].set_title(f'{model_name}')
        axes_summary[i].set_xlabel('SHAP value', fontsize=8)
        # Increase font size for x-axis labels
        plt.xticks(fontsize=8)
        plt.yticks(fontsize=8)

        # 特征bar图
        plt.sca(axes_summary_bar[i])
        shap.summary_plot(shap_values, X_test, show=False, plot_type="bar")
        axes_summary_bar[i].set_title(f'{model_name}')
        axes_summary_bar[i].set_xlabel('mean SHAP', fontsize=8)
        # Increase font size for x-axis labels
        plt.xticks(fontsize=8)
        plt.yticks(fontsize=8)

    fig_force.savefig('force_plots.png')
    fig_waterfall.savefig('waterfall_plots.png')
    fig_summary.savefig('summary_plots.png')
    fig_summary_bar.savefig('summary_plots_bar.png')

    plt.show()
    return shap_values_dict


if __name__ == '__main__':
    # 载入数据集
    x_train = pd.read_excel('2024数据整合_new.xlsx', sheet_name='x_train', header=0, engine='openpyxl')
    y_train = pd.read_excel('2024数据整合_new.xlsx', sheet_name='y_train', header=0, engine='openpyxl')
    x_test = pd.read_excel('2024数据整合_new.xlsx', sheet_name='x_test', header=0, engine='openpyxl')
    y_test = pd.read_excel('2024数据整合_new.xlsx', sheet_name='y_test', header=0, engine='openpyxl')

    # 移除索引列
    x_train = x_train.drop(columns=['Original_Index'])
    x_test = x_test.drop(columns=['Original_Index'])
    y_train_values = y_train.drop(columns=['Original_Index'])
    y_test_values = y_test.drop(columns=['Original_Index'])

    x_train_values = x_train.values
    y_train_values = np.array(y_train_values)  # ensure it is 1D for sklearn
    x_test_values = x_test.values
    y_test_values = np.array(y_test_values)

    models = init_models()
    Y_train_preds, Y_test_preds = train_and_evaluate_all(models, x_train_values, y_train_values, x_test_values)

    Y_preds = {}
    Y_true = {}

    for model_name in Y_train_preds.keys():
        Y_preds[model_name] = np.concatenate((Y_train_preds[model_name], Y_test_preds[model_name]))
    Y_true = np.concatenate((y_train_values, y_test_values))

    # 内部验证
    # 进行五折交叉验证并可视化结果
    print('Cross-validation')
    cross_validate_and_plot(models, np.concatenate((x_train_values, x_test_values)),
                            np.concatenate((y_train_values, y_test_values)))

    # 外部验证
    print('test')
    plot_all_predictions(y_test_values, Y_test_preds)
    plot_bland_altman(y_test_values, Y_test_preds)
    evaluation(y_test_values, Y_test_preds)
    plot_boxplot(y_test_values, Y_test_preds)

    print('train')
    plot_all_predictions(y_train_values, Y_train_preds)
    plot_bland_altman(y_train_values, Y_train_preds)
    evaluation(y_train_values, Y_train_preds)

    save_all_predictions_to_excel('predictions.xlsx', '0606  2024数据整合_训练(2).xlsx', '筛选0620', y_train, Y_train_preds, y_test, Y_test_preds)
    save_test_predictions_to_excel('predictions.xlsx', '验证组.xlsx', 'Sheet1', y_test, Y_test_preds)

    plot_jointgrid_multiple(y_train_values, Y_train_preds, y_test_values, Y_test_preds, models)

    shap_values_dict = generate_shap_plots(models, X_train, Y_train, X_test, Y_test)
    feature_names = pd.read_excel(file_name, sheet_name='x_train').drop(columns=['Original_Index']).columns  # 特征名称列表
    plot_shap_summary(shap_values_dict, feature_names)
